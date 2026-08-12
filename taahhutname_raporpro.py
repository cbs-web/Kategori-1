"""RaporPro ile aynı şablondan K-1 taahhütnamesi üretir.

Bu modül Word üzerinde görünümü yeniden kurmaz. RaporPro'nun doğrulanmış Excel
şablonunu doldurur ve Microsoft Excel üzerinden PDF'e aktarır. Böylece iki
programın yazı, hücre, kenarlık ve sayfa ölçeği aynı kalır.
"""

from __future__ import annotations

import math
import os
import re
import tempfile
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


TAAHHUT_SHEET = "tahhütname"
APP_DIR = os.path.dirname(os.path.abspath(__file__))
BUILTIN_TEMPLATE_PATH = os.path.join(
    APP_DIR,
    "ornek_sablonlar",
    "taahhutname",
    "taahhutname_base.xlsx",
)

TAAHHUT_METNI = (
    " Yukarıdaki bilgilere sahip projenin müellifliğini üstlenmemde 6235 sayılı Türk Mühendis ve "
    "Mimar Odaları Birliği Kanunu, 3194 sayılı İmar Kanunu ve ilgili mevzuat kapsamında süreli veya "
    "süresiz olarak mesleki faaliyet haklarımda herhangi bir kısıtlılık bulunmadığını ve odama "
    "üyeliğimin devam ettiğini taahhüt ederim.\n\u00a0\n"
    "Yukarıdaki bilgilere sahip yapıya ilişkin hazırlanacak tüm projelerde, 3194 sayılı Kanun ve "
    "deprem, yangın,enerji verimliliği, asansör gibi ilgili tüm mevzuat hükümlerini eksiksiz "
    "uygulayacağımı taahhüt ederim"
)

_SAYI_DESENI = re.compile(
    r"^[+-]?(?:\d+(?:[.,]\d*)?|[.,]\d+)(?:[eE][+-]?\d+)?$"
)


def _sayisal_metin_mi(metin):
    aday = str(metin).strip()
    if not _SAYI_DESENI.fullmatch(aday):
        return False
    try:
        return math.isfinite(float(aday.replace(",", ".")))
    except (TypeError, ValueError):
        return False


def _excel_hucre_degeri(deger):
    """Kullanıcı metninin Excel formülü olarak çalışmasını engelle."""
    if not isinstance(deger, str) or not deger:
        return deger
    aday = deger.lstrip()
    if not aday:
        return deger
    tehlikeli = aday[0] in ("=", "+", "@", "\t", "\r", "\n")
    tehlikeli = tehlikeli or (
        aday.startswith("-") and aday != "-" and not _sayisal_metin_mi(aday)
    )
    if tehlikeli and not _sayisal_metin_mi(aday):
        return "'" + deger
    return deger


def _set(ws, hucre, deger):
    ws[hucre] = _excel_hucre_degeri(deger)


def _temiz_dosya_adi(deger, varsayilan="Proje"):
    metin = str(deger or varsayilan).strip()
    metin = re.sub(r"[^\w\-.]+", "_", metin, flags=re.UNICODE).strip("._")
    return metin[:80] or varsayilan


def taahhutname_dosya_adi(uretici, tur, uzanti):
    rol = "Jeofizik" if tur == "jeofizik" else "Jeoloji"
    sahip = _temiz_dosya_adi(uretici.proje_deger("PROJE_ADI", "Proje"))
    uzanti = uzanti if str(uzanti).startswith(".") else f".{uzanti}"
    return f"{sahip}_Taahhutname_{rol}{uzanti}"


def _context(uretici, tur):
    tur = "jeofizik" if tur == "jeofizik" else "jeoloji"
    profil = uretici.taahhut_word_muhendis_bilgileri(tur)
    il = uretici.taahhut_yazim_duzeni(uretici.proje_deger("IL", ""))
    ilce = uretici.taahhut_yazim_duzeni(uretici.proje_deger("ILCE", ""))
    adres = uretici.taahhut_yapi_adresi()
    return {
        "tur": tur,
        "profil": profil,
        "il_ilce": " / ".join(parca for parca in (il, ilce) if parca),
        "ilgili_idare": uretici.taahhut_ilgili_idare(),
        "pafta_ada_parsel": uretici.taahhut_pafta_ada_parsel_metni(),
        "yapi_adresi": adres,
        "yapi_sahibi": uretici.taahhut_proje_adi_yazim_duzeni(),
        "yapi_sahibi_adresi": adres,
        "proje_turu": "ZEMİN ETÜDÜ RAPORU",
        "tarih": uretici.taahhut_tarihi(),
    }


def _calisma_kitabini_olustur(uretici, tur):
    if not os.path.isfile(BUILTIN_TEMPLATE_PATH):
        raise FileNotFoundError(
            f"RaporPro taahhütname şablonu bulunamadı: {BUILTIN_TEMPLATE_PATH}"
        )

    wb = load_workbook(BUILTIN_TEMPLATE_PATH)
    ws = wb.active
    ws.title = TAAHHUT_SHEET
    ws.sheet_view.showGridLines = False
    ctx = _context(uretici, tur)
    profil = ctx["profil"]

    deger_yazisi = Font(name="Aptos Narrow", size=11)
    imza_yazisi = Font(name="Times New Roman", size=11)
    sol = Alignment(horizontal="left")
    orta = Alignment(horizontal="center", vertical="center")

    degerler = {
        "C4": profil["sicil"],
        "C5": profil["unvan"],
        "C6": profil["adres"],
        "C7": profil["telefon"],
        "D11": ctx["il_ilce"],
        "D12": ctx["ilgili_idare"],
        "D13": ctx["pafta_ada_parsel"],
        "D14": ctx["yapi_adresi"],
        "D15": ctx["yapi_sahibi"],
        "D16": ctx["yapi_sahibi_adresi"],
        "D17": ctx["proje_turu"],
        "F30": ctx["tarih"],
        "F31": profil["ad"],
        "F32": profil["imza_unvan"],
    }
    for hucre, deger in degerler.items():
        _set(ws, hucre, deger)
        ws[hucre].font = imza_yazisi if hucre in {"F30", "F31", "F32"} else deger_yazisi
        ws[hucre].alignment = orta if hucre in {"F30", "F31", "F32"} else sol

    _set(ws, "A20", TAAHHUT_METNI)
    ws["A20"].font = deger_yazisi
    ws["A20"].alignment = Alignment(horizontal="left", wrap_text=True)

    for sayfa in wb.worksheets:
        sayfa.sheet_state = "hidden"
    ws.sheet_state = "visible"
    wb.active = wb.sheetnames.index(TAAHHUT_SHEET)
    ws.print_area = "A1:I47"
    ws.freeze_panes = "A1"
    for sutun in range(10, 19):
        ws.column_dimensions[get_column_letter(sutun)].hidden = True
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass
    wb.properties.creator = "K-1 / RaporPro"
    wb.properties.lastModifiedBy = "K-1 / RaporPro"
    return wb


def _xlsx_referans_ayarlarini_koru(yol):
    """RaporPro şablonunun Excel sütun/font varsayılanlarını koru."""
    gecici = f"{yol}.tmp"
    referans_sutunlar = (
        '<cols><col width="12.44140625" customWidth="1" style="43" min="1" max="1"/>'
        '<col width="6.33203125" customWidth="1" style="43" min="2" max="2"/>'
        '<col width="9.44140625" customWidth="1" style="43" min="5" max="5"/>'
        '<col width="10" customWidth="1" style="43" min="6" max="6"/>'
        '<col width="13.88671875" customWidth="1" style="43" min="8" max="8"/>'
        '<col width="8.6640625" customWidth="1" style="43" min="9" max="9"/>'
        + "".join(
            f'<col hidden="1" width="13" customWidth="1" style="43" min="{i}" max="{i}"/>'
            for i in range(10, 19)
        )
        + "</cols>"
    )
    try:
        with ZipFile(yol, "r") as zin:
            girdiler = [(item, zin.read(item.filename)) for item in zin.infolist()]
        stil_xml = next(
            (veri.decode("utf-8") for item, veri in girdiler if item.filename == "xl/styles.xml"),
            "",
        )
        referans_sutun_kullan = "Aptos Narrow" in stil_xml
        with ZipFile(gecici, "w", ZIP_DEFLATED) as zout:
            for item, veri in girdiler:
                if item.filename == "xl/styles.xml":
                    metin = veri.decode("utf-8")
                    metin = metin.replace(
                        '<name val="Calibri"/>',
                        '<name val="Aptos Narrow"/><charset val="162"/>',
                        1,
                    )
                    veri = metin.encode("utf-8")
                elif item.filename == "xl/worksheets/sheet1.xml" and referans_sutun_kullan:
                    metin = veri.decode("utf-8")
                    bas = metin.find("<cols>")
                    son = metin.find("</cols>")
                    if bas >= 0 and son >= 0 and 'style="43"' in metin[bas:son]:
                        metin = f"{metin[:bas]}{referans_sutunlar}{metin[son + len('</cols>'):]}"
                        veri = metin.encode("utf-8")
                zout.writestr(item, veri)
        os.replace(gecici, yol)
    finally:
        if os.path.exists(gecici):
            try:
                os.remove(gecici)
            except OSError:
                pass


def taahhut_xlsx_kaydet(uretici, tur, hedef_yol):
    hedef_yol = os.path.abspath(hedef_yol)
    klasor = os.path.dirname(hedef_yol)
    os.makedirs(klasor, exist_ok=True)
    wb = _calisma_kitabini_olustur(uretici, tur)
    gecici_yol = ""
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{os.path.basename(hedef_yol)}.",
            suffix=".tmp.xlsx",
            dir=klasor,
            delete=False,
        ) as dosya:
            gecici_yol = dosya.name
        wb.save(gecici_yol)
        _xlsx_referans_ayarlarini_koru(gecici_yol)
        kontrol = load_workbook(gecici_yol, read_only=True)
        kontrol.close()
        os.replace(gecici_yol, hedef_yol)
        gecici_yol = ""
        return hedef_yol
    finally:
        try:
            wb.close()
        except Exception:
            pass
        if gecici_yol:
            try:
                os.remove(gecici_yol)
            except OSError:
                pass


def _pdf_dogrulama_hatasi(yol):
    if not os.path.isfile(yol) or os.path.getsize(yol) <= 0:
        return "PDF dosyası oluşmadı."
    try:
        okuyucu = PdfReader(yol)
        if len(okuyucu.pages) != 1:
            return f"Taahhütname PDF'i bir sayfa olmalı; bulunan: {len(okuyucu.pages)}."
    except Exception as exc:
        return f"PDF okunamadı: {exc}"
    return ""


def _com_ozelligini_ayarla(nesne, ad, deger):
    try:
        setattr(nesne, ad, deger)
    except Exception:
        pass


def _excel_pdf_bir_kez(xlsx_yolu, pdf_yolu):
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        return False, f"Microsoft Excel PDF aktarımı için pywin32 bulunamadı: {exc}"

    excel = None
    kitap = None
    com_hazir = False
    aktarim_hatasi = None
    try:
        pythoncom.CoInitialize()
        com_hazir = True
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        _com_ozelligini_ayarla(excel, "AutomationSecurity", 3)
        _com_ozelligini_ayarla(excel, "AskToUpdateLinks", False)
        _com_ozelligini_ayarla(excel, "EnableEvents", False)
        kitap = excel.Workbooks.Open(
            os.path.abspath(xlsx_yolu),
            UpdateLinks=0,
            ReadOnly=True,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
            Notify=False,
        )
        sayfa = kitap.Worksheets(TAAHHUT_SHEET)
        sayfa.ExportAsFixedFormat(0, os.path.abspath(pdf_yolu))
    except Exception as exc:
        aktarim_hatasi = exc
    finally:
        if kitap is not None:
            try:
                kitap.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        if com_hazir:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    dogrulama = _pdf_dogrulama_hatasi(pdf_yolu)
    if not dogrulama:
        return True, ""
    if aktarim_hatasi is not None:
        return False, f"{aktarim_hatasi}; {dogrulama}"
    return False, dogrulama


def taahhut_xlsx_pdfye_cevir(xlsx_yolu, pdf_yolu):
    """Excel aktarımını yeni COM oturumuyla yap; bir kez güvenli tekrar dene."""
    xlsx_yolu = os.path.abspath(xlsx_yolu)
    pdf_yolu = os.path.abspath(pdf_yolu)
    hatalar = []
    for _deneme in range(2):
        if os.path.exists(pdf_yolu):
            try:
                os.remove(pdf_yolu)
            except OSError:
                pass
        basarili, hata = _excel_pdf_bir_kez(xlsx_yolu, pdf_yolu)
        if basarili:
            return True, ""
        hatalar.append(hata)
    return False, " | ".join(hata for hata in hatalar if hata)

